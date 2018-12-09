#!/usr/bin/env python
# -*- coding: utf-8 -*-
__author__ = 'orleven'

import time
import os
import re
import sys

from openpyxl import Workbook
from urllib import parse
from lib.data import logger
from lib.config import load_conf
from lib.config import init_conf
from lib.update import update_program

def banner():
    banner = """
        _____         _     ___  ___
       /  ___|       | |    |  \/  |
       \ `--.  _   _ | |__  | .  . |  ___   _ __
        `--. \| | | || '_ \ | |\/| | / _ \ | '_ \\
       /\__/ /| |_| || |_) || |  | || (_) || | | |  version: 1.0
       \____/  \__,_||_.__/ \_|  |_/ \___/ |_| |_|  author: @orleven

"""
    print(banner)

def check_domain(url):
    """
    check domain
    valid domain: http://example.com or https://example.com or example.com
    :param url:
    :return:
    """
    url = url.replace('\r','').replace('\n','').strip()

    pattern = re.compile("^(http://|https://)?[a-zA-Z0-9]+([\-\.]{1}[a-zA-Z0-9]+)*\.[a-zA-Z]{2,}$")
    if not pattern.match(url):
        return False

    if not url.startswith('http://') and not url.startswith('https://'):
        url = 'http://' + url

    return parse.urlparse(url)


def get_safe_ex_string(ex, encoding=None):
    """
    Safe way how to get the proper exception represtation as a string
    (Note: errors to be avoided: 1) "%s" % Exception(u'\u0161') and 2) "%s" % str(Exception(u'\u0161'))

    >>> get_safe_ex_string(Exception('foobar'))
    u'foobar'
    """

    retVal = ex

    if getattr(ex, "message", None):
        retVal = ex.message
    elif getattr(ex, "msg", None):
        retVal = ex.msg
    return retVal
    # return getUnicode(retVal or "", encoding=encoding).strip()

def config_parser():
    path = os.path.join(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),'submon.conf'))
    if not os.path.exists(path):
        init_conf(path)
    load_conf(path)


def check_update(args):
    if args.update:
        update_program()
        sys.exit(0)

def tocsv(datalines,path,file):
    filename = os.path.join(path,file)
    logger.info('Export to %s...' % (filename))
    book = Workbook()
    ws = book.active
    i = 1
    titleList = []
    for line in datalines:
        i = i + 1
        for key in line:
            if key not in titleList:
                titleList.append(key)
                ws.cell(row=1, column=len(titleList)).value = key
            try:
                if isinstance(line[key], int) or isinstance(line[key], str):
                    ws.cell(row=i, column=titleList.index(key) + 1).value = line[key]
                elif isinstance(line[key], list):
                    ws.cell(row=i, column=titleList.index(key) + 1).value = str(line[key])
                elif isinstance(line[key], dict):
                    ws.cell(row=i, column=titleList.index(key) + 1).value = str(line[key])
                else:
                    ws.cell(row=i, column=titleList.index(key) + 1).value = "Types of printing are not supported."
            except:
                ws.cell(row=i, column=titleList.index(key) + 1).value = "Some error."
    book.save(filename)
    logger.sysinfo('Exported to %s successful!' % (filename))

#
